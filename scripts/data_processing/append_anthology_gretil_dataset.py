#!/usr/bin/env python3
"""
Append Sanskrit verses from Bhartṛhari, Mahāsubhāṣitasaṅgraha, and Subhāṣitaratnakoṣa
to build_gradil_dataset.xlsx.

Format: [verse_id, shloka] (Devanagari)
REFINED: Removes author names, Citations, and header junk.
"""

import re
import openpyxl
from indic_transliteration import sanscript
from indic_transliteration.sanscript import transliterate

XLSX_PATH = "/Users/suhritghimire/Downloads/NavaRasaBank-main/build_gradil_dataset.xlsx"

def strip_tags(s):
    s = re.sub(r"<[^>]+>", "", s)
    s = s.replace("&lt;", "<").replace("&gt;", ">").replace("&amp;", "&")
    return s.strip()

def clean_verse(text):
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"[\*]", "", text)
    # Remove markers like $ & % used as line breaks in SRK
    text = re.sub(r"[\$&\%\/]", "।", text)
    # Collapse multiple dandas and whitespace
    text = re.sub(r"।+", "।", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def translit_and_clean(text):
    dev = transliterate(text, sanscript.IAST, sanscript.DEVANAGARI)
    dev = dev.replace("||", "॥").replace("|", "।")
    # Fix the double danda if transliteration created it
    dev = dev.replace("।।", "॥")
    # Final cleanup of common artifacts
    dev = re.sub(r"\s+", " ", dev).strip()
    return dev

def is_junk(line):
    # Citation patterns like "(Skmsa...)" or "[p. 123]"
    if re.search(r"^\(.*\d+.*\)", line): return True
    if re.search(r"^\[p\.\s*\d+\]", line): return True
    # Author names ending with -- or names like "śīlāmbhaḥ..."
    if line.endswith("--"): return True
    # GRETIL Metadata
    if any(k in line for k in ["GRETIL", "description:", "long a", "multibyte", "Copyright", "Input", "Asko Parpola", "browser's VIEW"]): return True
    # Section titles like "1. sugata-vrajyā"
    if re.match(r"^\d+\.\s+[a-z]", line): return True
    return False

all_verses = []

# --- 1. Bhartṛhari (bharst_u.htm) ---
print("Processing Bhartṛhari...")
with open("/tmp/bharst_u.htm", "r", encoding="utf-8", errors="replace") as f:
    # Split by <BR> first
    text = re.sub(r"<[Bb][Rr]\s*/?>", "\n", f.read())
    lines = [strip_tags(l) for l in text.splitlines() if strip_tags(l)]

buf = []
pattern = re.compile(r"\|\|\s*(BharSt_[\d.]+)\s*\|\|")
for line in lines:
    if is_junk(line): continue
    match = pattern.search(line)
    if match:
        v_id = match.group(1).replace("BharSt", "Bhartri")
        content = pattern.sub("", line).strip()
        if content: buf.append(content)
        # Final check for first verse start
        v_text = " ".join(buf)
        if v_id == "Bhartri_1.1":
            m_start = v_text.find("dik-kālādy")
            if m_start != -1: v_text = v_text[m_start:]
        
        all_verses.append((v_id, translit_and_clean(clean_verse(v_text))))
        buf = []
    else:
        buf.append(line)

# --- 2. Mahāsubhāṣitasaṅgraha (msubhs_u.htm) ---
print("Processing Mahāsubhāṣitasaṅgraha...")
with open("/tmp/msubhs_u.htm", "r", encoding="utf-8", errors="replace") as f:
    text = f.read()
    text = re.sub(r"<[Bb][Rr]\s*/?>", "\n", text)
    lines = [strip_tags(l) for l in text.splitlines() if strip_tags(l)]

msubh_groups = {}
mss_pattern = re.compile(r"MSS_(\d+)-(\d+)")
for line in lines:
    match = mss_pattern.search(line)
    if match:
        v_num = match.group(1)
        content = mss_pattern.sub("", line).strip()
        if v_num not in msubh_groups: msubh_groups[v_num] = []
        if content: msubh_groups[v_num].append(content)

for v_num in sorted(msubh_groups.keys(), key=int):
    full_text = " ".join(msubh_groups[v_num])
    if full_text.strip():
        all_verses.append((f"MSubh_{int(v_num):05d}", translit_and_clean(clean_verse(full_text))))

# --- 3. Subhāṣitaratnakoṣa (vidsrgau.htm) ---
print("Processing Subhāṣitaratnakoṣa...")
with open("/tmp/vidsrgau.htm", "r", encoding="utf-8", errors="replace") as f:
    text = f.read()
    text = re.sub(r"<[Bb][Rr]\s*/?>", "\n", text)
    lines = [strip_tags(l) for l in text.splitlines() if strip_tags(l)]

buf = []
pattern = re.compile(r"//\s*(VidSrk_[\d.]+)\b")
for line in lines:
    if is_junk(line): continue
    match = pattern.search(line)
    if match:
        v_id = match.group(1).replace("VidSrk", "SRK")
        content = pattern.sub("", line).strip()
        content = re.sub(r"\*\(\d+\)\s*$", "", content).strip()
        if content: buf.append(content)
        all_verses.append((v_id, translit_and_clean(clean_verse(" ".join(buf)))))
        buf = []
    else:
        buf.append(line)

# --- SAVE ---
print(f"Total anthology verses: {len(all_verses)}")
wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.active

# Clear previous anthology data (keep first 3136 rows)
if ws.max_row > 3136:
    print(f"Deleting {ws.max_row - 3136} existing anthology rows...")
    ws.delete_rows(3137, ws.max_row - 3136)

for v_id, shloka in all_verses:
    ws.append([v_id, shloka])

wb.save(XLSX_PATH)
print("Saved refined anthology verses.")
