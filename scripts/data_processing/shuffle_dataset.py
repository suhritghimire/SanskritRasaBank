import openpyxl
import random

XLSX_PATH = "/Users/suhritghimire/Downloads/NavaRasaBank-main/build_gradil_dataset.xlsx"

def shuffle_dataset():
    print("Loading Excel file...")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    
    # Get all rows
    data = list(ws.iter_rows(values_only=True))
    if not data:
        print("Dataset is empty.")
        return
        
    header = data[0]
    verses = data[1:]
    
    print(f"Shuffling {len(verses)} verses...")
    # Shuffle in-place
    random.shuffle(verses)
    
    print("Writing shuffled data back...")
    # Clear and rewrite
    ws.delete_rows(1, ws.max_row)
    ws.append(header)
    for row in verses:
        ws.append(row)
        
    # Re-apply bold to header
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        
    wb.save(XLSX_PATH)
    print("Dataset shuffled successfully.")

if __name__ == "__main__":
    shuffle_dataset()
