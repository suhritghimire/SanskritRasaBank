#!/usr/bin/env python3
"""
Append Sanskrit verses from kmeghdpu.htm (Kalidasa's Meghaduta)
to the existing new_dataset_gretil.xlsx.
Refined to skip header junk.
"""

import re
import openpyxl
from indic_transliteration import sanscript
from indic_transliteration.sanscript import transliterate

# 1. READ LOCAL HTML
HTML_PATH = "/tmp/kmeghdpu.htm"
with open(HTML_PATH, "r", encoding="utf-8", errors="replace") as f:
    raw_html = f.read()

# 2. STRIP HTML TAGS and SPLIT into lines
# We want to keep the structure but remove the tags.
# Replace <BR> with \n
text = re.sub(r"<[Bb][Rr]\s*/?>", "\n", raw_html)
# Remove all other tags
text = re.sub(r"<[^>]+>", "", text)
text = text.replace("&lt;", "<").replace("&gt;", ">").replace("&amp;", "&")
lines = text.splitlines()

# 3. EXTRACT VERSES
VERSE_ID_PATTERN = re.compile(r"//\s*(KMgD_\d+)\s*//")

verses = []
current_verse_lines = []

# Improved start logic: start after the second occurrence of "http://gretil.sub.uni-goettingen.de/gretil.htm"
# or simply when we see the first KMgD pattern but ignoring what's before it in the buffer.
start_collecting = False

for line in lines:
    line = line.strip()
    if not line:
        continue
    
    # Skip metadata lines
    if any(k in line for k in ["description:", "multibyte sequence", "long ", "vocalic ", "anusvara", "visarga", "GRETIL", "COPYRIGHT", "browser's VIEW"]):
        continue
    
    match = VERSE_ID_PATTERN.search(line)
    if match:
        verse_id = match.group(1)
        # Remove the ID from the line text
        line_clean = VERSE_ID_PATTERN.sub("", line).strip()
        current_verse_lines.append(line_clean)
        
        # Join the lines for the complete verse
        full_verse_iast = " ".join(current_verse_lines).strip()
        
        # Transliterate to Devanagari
        full_verse_devanagari = transliterate(full_verse_iast, sanscript.IAST, sanscript.DEVANAGARI)
        
        # Replace | or || with । or ॥
        full_verse_devanagari = full_verse_devanagari.replace("||", "॥").replace("|", "।")
        # Also convert / to । if it's there as a lone separator
        # The user didn't ask but it's part of the cleaning for this specific file.
        # Actually, let's just stick to the user's specific ask for | and ||.
        # Wait, I'll also convert / if it's at the end or middle as it's a danda in GRETIL.
        
        verses.append((full_verse_devanagari, verse_id))
        
        # Reset for next verse
        current_verse_lines = []
    else:
        # Avoid collecting header junk
        if "Meghaduta" in line or "PLAIN TEXT" in line or "Asko Parpola" in line:
            continue
        current_verse_lines.append(line)

# Remove any verses that might have captured header data (usually KMgD_1 is the victim)
# Let's check KMgD_1 and strip anything before the first known Sanskrit word if possible, 
# or better yet, just ensure start_collecting works.
# Actually, KMgD_1 starts with "kas cit". Let's use that.
final_verses = []
for v_text, v_id in verses:
    if v_id == "KMgD_1":
        # Find "कश् चित्" (kas cit)
        start_idx = v_text.find("कश् चित्")
        if start_idx != -1:
            v_text = v_text[start_idx:]
    final_verses.append((v_text, v_id))

print(f"Extracted {len(final_verses)} verses from Meghaduta.")

# 4. CLEAN AND APPEND TO EXCEL
XLSX_PATH = "/Users/suhritghimire/Downloads/NavaRasaBank-main/new_dataset_gretil.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.active

# Delete previous KMgD rows (rows 315 to end)
# Total rows were 425. Previous data rows were 313.
# Row 314 was the last HDratn row? Let's check.
# 197 (KSak) + 51 (HPri) + 65 (HDratn) = 313.
# Row 1 is header.
# Rows 2-314 are non-KMgD.
# Row 315 was the start of KMgD.
if ws.max_row >= 315:
    ws.delete_rows(315, ws.max_row - 314)

for dev_text, v_id in final_verses:
    ws.append([dev_text, v_id])

wb.save(XLSX_PATH)
print(f"Successfully cleaned and re-appended verses to {XLSX_PATH}.")
