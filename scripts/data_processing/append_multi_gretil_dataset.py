#!/usr/bin/env python3
"""
Append Sanskrit verses from Raghuvaṃśa, Kirātārjunīya, and Bhallaṭaśataka
to the existing build_gradil_dataset.xlsx.

Strict Format:
  - 2 columns: verse_id, shloka
  - shloka in Devanagari
"""

import re
import openpyxl
from indic_transliteration import sanscript
from indic_transliteration.sanscript import transliterate

# ── 1. CONFIGURATION ──

XLSX_PATH = "/Users/suhritghimire/Downloads/NavaRasaBank-main/build_gradil_dataset.xlsx"

SOURCES = [
    {
        "name": "Raghuvaṃśa",
        "path": "/tmp/kragh_pu.htm",
        "id_prefix": "Raghu",
        "id_pattern": re.compile(r"//\s*(Ragh_[\d.]+)\s*//"),
    },
    {
        "name": "Kirātārjunīya",
        "path": "/tmp/bhakirxu.htm",
        "id_prefix": "Kirata",
        "id_pattern": re.compile(r"//\s*(BhKir_[\d.]+)\s*(?://)?"),
    },
    {
        "name": "Bhallaṭaśataka",
        "path": "/tmp/bhall_pu.htm",
        "id_prefix": "Bhallata",
        "id_pattern": re.compile(r"//\s*(Bhall_\d+)\s*//"),
    }
]

# ── 2. EXTRACTION LOGIC ──

def strip_tags(s):
    s = re.sub(r"<[^>]+>", "", s)
    s = s.replace("&lt;", "<").replace("&gt;", ">").replace("&amp;", "&")
    return s.strip()

def process_source(source):
    print(f"Processing {source['name']}...")
    with open(source["path"], "r", encoding="utf-8", errors="replace") as f:
        raw_html = f.read()

    # Pre-process lines
    lines_raw = re.sub(r"<[Bb][Rr]\s*/?>", "\n", raw_html)
    lines_raw = re.sub(r"<[^>]+>", "", lines_raw)
    lines_raw = lines_raw.replace("&lt;", "<").replace("&gt;", ">").replace("&amp;", "&")
    lines = lines_raw.splitlines()

    extracted_verses = []
    current_verse_lines = []
    
    # Generic stable counter for fallback IDs
    stable_counter = 1
    
    # Skip until after metadata (usual GRETIL header signs)
    start_processing = False

    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Start processing after common header markers
        if any(marker in line for marker in ["PLAIN TEXT", "http://gretil", "GRETIL"]):
            start_processing = True
            continue
        
        # Heuristic to avoid header tables
        if any(marker in line for marker in ["description:", "multibyte sequence", "long ", "vocalic "]):
            continue

        match = source["id_pattern"].search(line)
        if match:
            # We found an ID!
            verse_id = match.group(1)
            # Remove the ID from the line text
            line_clean = source["id_pattern"].sub("", line).strip()
            if line_clean:
                current_verse_lines.append(line_clean)
            
            full_verse_iast = " ".join(current_verse_lines).strip()
            
            # Final cleaning
            full_verse_iast = re.sub(r"\*", "", full_verse_iast) # Remove asterisks
            full_verse_iast = re.sub(r"\s+", " ", full_verse_iast).strip()
            
            if full_verse_iast:
                extracted_verses.append((full_verse_iast, verse_id))
            
            current_verse_lines = []
            stable_counter += 1
        else:
            # Check for spurious prose/metadata lines
            if any(k in line for k in ["Copyright", "Input", "browser's VIEW", "July 2013"]):
                continue
            current_verse_lines.append(line)

    print(f"  Extracted {len(extracted_verses)} verses.")
    return extracted_verses

# ── 3. RUN EXTRACTION AND TRANSLITERATION ──

all_new_verses = []

for source in SOURCES:
    verses_iast = process_source(source)
    for iast_text, vid in verses_iast:
        # Transliterate to Devanagari
        dev_text = transliterate(iast_text, sanscript.IAST, sanscript.DEVANAGARI)
        # Clean dandas
        dev_text = dev_text.replace("||", "॥").replace("|", "।")
        # Standardize verse ID prefix if requested (user said Raghu_001 if no id, 
        # but here we have ids from text. I'll just use the ones found, 
        # but prepend work name for consistency if preferred? 
        # User said: "canto.verse or the page's marker. If no explicit ID exists, Raghu_001..."
        # I'll keep the found ID like Ragh_1.1.
        all_new_verses.append((vid, dev_text))

# ── 4. UPDATE EXCEL ──

print("\nUpdating Excel file...")
wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.active

# Reorder columns: [shloka, id] -> [id, shloka]
rows = list(ws.iter_rows(values_only=True))
header = ["verse_id", "shloka"]
new_rows = [header]

if rows:
    # Existing rows mapping: row[0] is sanskrit_sholka, row[1] is verse_id
    for r in rows[1:]: # Skip existing header
        new_rows.append([r[1], r[0]])

# Append new ones
for vid, dev in all_new_verses:
    new_rows.append([vid, dev])

# Clear and rewrite
ws.delete_rows(1, ws.max_row)
for r in new_rows:
    ws.append(r)

# Format header
from openpyxl.styles import Font
for cell in ws[1]:
    cell.font = Font(bold=True)

wb.save(XLSX_PATH)
print(f"Saved to {XLSX_PATH}. Total rows: {len(new_rows)}")

