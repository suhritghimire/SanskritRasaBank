#!/usr/bin/env python3
"""
Final Refined SRK cleaning.
- Matches specific author/attribution patterns at the very start.
- Removes citation-only parentheses (digits or keywords) while preserving others.
- Normalizes formatting.
"""

import re
import openpyxl

XLSX_PATH = "/Users/suhritghimire/Downloads/NavaRasaBank-main/build_gradil_dataset.xlsx"

CIT_KEYWORDS = ['स्स्', 'वैस्', 'स्क्म्स', 'स्व्सु', 'स्प्द्शा', 'स्क्स', 'Skmsa', 'Skms', 'स्भ्सु', 'स्भ्क']

def is_citation_paren(p_text):
    # CITATION RULE: parentheses (...) that contain digits (०-९/0-9) OR citation keywords
    if re.search(r'[०-९0-9]', p_text):
        return True
    for kw in CIT_KEYWORDS:
        if kw in p_text:
            return True
    return False

def clean_srk_text(text):
    if not text:
        return ""
        
    # 1. Remove leading author/attribution stub: <NAME>स्य । or कस्यचित् ।
    # Rules: starts at the beginning, ends at the first ।
    # Must be a single word (no spaces) ending in स्य or is कस्यचित् ।
    # Pattern: ^([^।\s]+(?:स्य|चित्) ।)\s*
    attr_match = re.match(r'^([^।\s]+(?:स्य|चित्) ।)\s*', text)
    if attr_match:
        text = text[len(attr_match.group(0)):].strip()

    # 2. Remove citation parentheses (...)
    # Logic: find all parentheses blocks and remove them if they match the citation rule
    # Using re.finditer to avoid replacing correct verse parens that might have identical inner text
    # However, replacing all occurrences of a specific string is usually safe if it's a unique citation.
    
    # Simple regex to find parentheses
    parens = re.findall(r'\([^)]+\)', text)
    for p in parens:
        inner = p[1:-1]
        if is_citation_paren(inner):
            # Escape for safety if using re.sub, but simple replace is fine here for exact matches
            text = text.replace(p, '')

    # 3. Normalize formatting
    # trimmed leading/trailing spaces
    text = text.strip()
    # remove any leading । left behind
    if text.startswith('।'):
        text = text[1:].strip()
    # collapse multiple spaces into one
    text = re.sub(r'\s+', ' ', text).strip()
    
    return text

def run_fix():
    print("Loading Excel...")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    
    header = ["verse_id", "shloka"]
    new_rows = []
    
    print("Repairing SRK rows...")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        v_id = str(row[0]) if row[0] else ""
        shloka = str(row[1]) if row[1] else ""
        
        if v_id.startswith("SRK_") or v_id.startswith("VidSrk_"):
            shloka = clean_srk_text(shloka)
            if not shloka: continue # skip empty
            
        new_rows.append([v_id, shloka])
        
    print("Updating file...")
    ws.delete_rows(1, ws.max_row)
    ws.append(header)
    for r in new_rows:
        ws.append(r)
        
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        
    wb.save(XLSX_PATH)
    print("SRK refinement complete.")

if __name__ == "__main__":
    run_fix()
