#!/usr/bin/env python3
"""
Append Sanskrit verses from hpriydau.htm (HPri = Harsha's Priyadarshika)
to the existing new_dataset_gretil.xlsx.

HPri format: verses span multiple plain HTML lines (no * prefix).
Last line of each verse contains // HPri_X.Y //
Verses separated from prose by blank <BR> lines.
"""

import re
import openpyxl
from openpyxl.styles import Font

# ── 1. Read raw HTML ─────────────────────────────────────
with open("/tmp/hpriydau.htm", encoding="utf-8", errors="replace") as f:
    raw_lines = f.read().splitlines()

print(f"Total raw lines: {len(raw_lines)}")

# ── Helper: strip ALL HTML tags and common entities ──────
def strip_tags(s):
    s = re.sub(r"<[^>]+>", "", s)
    s = s.replace("&lt;","<").replace("&gt;",">").replace("&amp;","&")
    s = s.replace("|", " ").strip()   # HPri uses | as a prose separator
    return s.strip()

# ── 2. Patterns ──────────────────────────────────────────
ID_PAT   = re.compile(r"//\s*(HPri_[\d.]+[a-z]?)\s*//")
# Prose signatures: lines that are clearly NOT verse content
PROSE_PAT = re.compile(
    r"(\([^)]+\)|sūtradhāraḥ|rājā|vidūṣaka|kañcukī|naṭī|iti|nepathya"
    r"|prathamo|dvitīyo|tṛtīyo|caturtho|pañcamo|ṣaṣṭho|saptamo"
    r"|praviśati|niṣkrāntaḥ|viṣkambhakaḥ|prastāvanā)",
    re.IGNORECASE
)

# ── 3. Extract verse blocks ───────────────────────────────
#
# Strategy:
#  - Walk raw lines
#  - Each raw line is converted to clean text
#  - Accumulate lines into a buffer
#  - When the ID pattern // HPri_... // appears, flush as a verse
#  - Clear buffer on blank lines appearing *before* any ID is found
#    (these separate prose from verse preamble)
#  - Skip lines that look like pure prose / stage directions

records  = []
buf      = []
cur_id   = ""

SKIP_PAT = re.compile(
    r"^\s*(\[p\.\s*\d+\]|http|GRETIL|Input|Copyright|THIS|Text converted"
    r"|Unless|description:|long [aiuAIU]|vocalic|retroflex|palatal|velar"
    r"|anusvara|visarga|For a comp|For further|---)",
    re.IGNORECASE
)

for raw_line in raw_lines:
    clean = strip_tags(raw_line).strip()

    # Skip editorial/metadata lines
    if SKIP_PAT.match(clean):
        continue

    # Blank line → flush only if no ID has been seen (orphan prose lines)
    if not clean:
        if buf and not cur_id:
            buf = []       # discard prose preamble without a verse ID
        continue

    # Check if this line carries the closing ID
    id_m = ID_PAT.search(clean)
    if id_m:
        cur_id = id_m.group(1)
        content = ID_PAT.sub("", clean).strip().strip(" /|.")
        if content:
            buf.append(content)
        # Flush the complete verse
        verse = " ".join(buf)
        verse = re.sub(r"\s+", " ", verse).strip()
        verse = verse.strip(" /|.")
        # Remove page refs
        verse = re.sub(r"\[p\.\s*\d+\]", "", verse).strip()
        if len(verse) > 8:
            records.append((cur_id, verse))
        buf    = []
        cur_id = ""
        continue

    # Is this a prose / stage direction line? Skip it.
    # But only if the buffer is empty (we haven't started a verse yet)
    if not buf and PROSE_PAT.search(clean):
        continue

    # If we already have verse content in buf, stop at prose lines
    if buf and PROSE_PAT.search(clean):
        buf = []   # abandon incomplete verse that got interrupted by prose
        continue

    # Plain content line — add to buffer
    clean = re.sub(r"\[p\.\s*\d+\]", "", clean).strip()
    if clean:
        buf.append(clean)

# tail flush
if buf and cur_id:
    verse = " ".join(buf)
    if len(verse) > 8:
        records.append((cur_id, verse))

print(f"Extracted {len(records)} HPri verses")

# ── 4. Transliterate IAST → Devanagari ───────────────────
def transliterate(text):
    try:
        from indic_transliteration import sanscript
        from indic_transliteration.sanscript import transliterate as tr
        return tr(text, sanscript.IAST, sanscript.DEVANAGARI)
    except Exception:
        pass
    TABLE = [
        ("ā","आ"),("Ā","आ"),("ī","ई"),("Ī","ई"),("ū","ऊ"),("Ū","ऊ"),
        ("ṝ","ॠ"),("ṛ","ऋ"),("Ṛ","ऋ"),("ḷ","ऌ"),("ḹ","ॡ"),
        ("ai","ऐ"),("au","औ"),("e","ए"),("o","ओ"),
        ("a","अ"),("i","इ"),("u","उ"),
        ("ṃ","ं"),("ḥ","ः"),
        ("kh","ख"),("gh","घ"),("ch","छ"),("jh","झ"),
        ("ṭh","ठ"),("ḍh","ढ"),("th","थ"),("dh","ध"),
        ("ph","फ"),("bh","भ"),
        ("k","क"),("g","ग"),("ṅ","ङ"),
        ("c","च"),("j","ज"),("ñ","ञ"),
        ("ṭ","ट"),("ḍ","ड"),("ṇ","ण"),
        ("t","त"),("d","द"),("n","न"),
        ("p","प"),("b","ब"),("m","म"),
        ("y","य"),("r","र"),("l","ल"),("v","व"),
        ("ś","श"),("Ś","श"),("ṣ","ष"),("Ṣ","ष"),
        ("s","स"),("h","ह"),("ḻ","ळ"),
        ("||","॥"),("|","।"),
    ]
    out, i = [], 0
    while i < len(text):
        for src, dst in TABLE:
            if text[i:i+len(src)] == src:
                out.append(dst); i += len(src); break
        else:
            out.append(text[i]); i += 1
    return "".join(out)

print("Transliterating …")
final = [(vid, transliterate(v)) for vid, v in records]
print(f"Done — {len(final)} rows to append")

# ── 5. Append to existing Excel file ─────────────────────
XLSX = "/Users/suhritghimire/Downloads/NavaRasaBank-main/new_dataset_gretil.xlsx"
wb   = openpyxl.load_workbook(XLSX)
ws   = wb.active

existing_rows = ws.max_row
print(f"Existing rows (incl header): {existing_rows}")

for vid, dev in final:
    ws.append([dev, vid])

wb.save(XLSX)

print(f"\n✅  Saved → {XLSX}")
print(f"   Previous data rows: {existing_rows - 1}")
print(f"   Appended HPri rows: {len(final)}")
print(f"   Total data rows now: {ws.max_row - 1}")

print("\nSample HPri rows appended:")
for vid, dev in final[:5]:
    print(f"  [{vid}]  {dev[:90]}")
