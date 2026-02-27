#!/usr/bin/env python3
"""
Repair KSS Vetāla verses by merging half-lines into full ślokas.
In-place repair of build_gradil_dataset.xlsx.
"""

import re
import openpyxl
from indic_transliteration import sanscript
from indic_transliteration.sanscript import transliterate

XLSX_PATH = "/Users/suhritghimire/Downloads/NavaRasaBank-main/build_gradil_dataset.xlsx"
HTML_PATH = "/tmp/soksvppu.htm"

def strip_tags(s):
    s = re.sub(r"<[^>]+>", "", s)
    s = s.replace("&lt;", "<").replace("&gt;", ">").replace("&amp;", "&")
    return s.strip()

def clean_and_translit(text):
    text = re.sub(r"[\*_]", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    text = text.replace('||', '॥').replace('|', '।')
    dev = transliterate(text, sanscript.IAST, sanscript.DEVANAGARI)
    dev = dev.replace('।।', '॥')
    return dev.strip()

def extract_merged_kss():
    with open(HTML_PATH, "r", encoding="utf-8", errors="replace") as f:
        text_content = re.sub(r"<[Bb][Rr]\s*/?>", "\n", f.read())
        lines = [strip_tags(l) for l in text_content.splitlines() if strip_tags(l)]

    # Pattern: // SoKss_12,8.x (Vet_y.z) //
    kss_pattern = re.compile(r"//\s*SoKss_[\d.,]+\s*\((Vet_([\d.]+))\)\s*//")
    extracted_parts = []
    for line in lines:
        if "GRETIL" in line or "description:" in line: continue
        match = kss_pattern.search(line)
        if match:
            v_id_raw = match.group(2) # e.g. 0.1
            content = kss_pattern.sub("", line).strip()
            extracted_parts.append({'id': v_id_raw, 'text': content})

    merged_verses = []
    # Merge pairs: (0.1, 0.2), (0.3, 0.4), etc.
    for i in range(0, len(extracted_parts) - 1, 2):
        p1 = extracted_parts[i]
        p2 = extracted_parts[i+1]
        
        # New ID: KSS_Vet_0_1-2
        # Strip the last part for the range
        prefix = ".".join(p1['id'].split('.')[:-1])
        start = p1['id'].split('.')[-1]
        end = p2['id'].split('.')[-1]
        v_id = f"KSS_Vet_{prefix.replace('.', '_')}_{start}-{end}"
        
        full_text = p1['text'] + " " + p2['text']
        merged_verses.append((v_id, clean_and_translit(full_text)))

    return merged_verses

def repair_kss_in_excel(new_verses):
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    
    old_data = list(ws.iter_rows(values_only=True))
    header = old_data[0] # assuming 3-column at this point
    
    cleaned_rows = []
    for row in old_data[1:]:
        v_id = str(row[0])
        if v_id.startswith("KSS_Vet_"):
            continue
        cleaned_rows.append(row)
        
    final_rows = [header]
    kss_inserted = False
    
    for row in cleaned_rows:
        v_id = str(row[0])
        # Insert before Tantra1 or at end
        if v_id.startswith("Tantra1") and not kss_inserted:
            for vid, txt in new_verses:
                final_rows.append([vid, "Kathāsaritsāgara (Vetala)", txt])
            kss_inserted = True
        final_rows.append(row)
        
    if not kss_inserted:
        for vid, txt in new_verses:
            final_rows.append([vid, "Kathāsaritsāgara (Vetala)", txt])

    ws.delete_rows(1, ws.max_row)
    for r in final_rows:
        ws.append(r)
    wb.save(XLSX_PATH)
    print(f"KSS Vetala repaired with {len(new_verses)} merged ślokas.")

if __name__ == "__main__":
    v = extract_merged_kss()
    repair_kss_in_excel(v)
