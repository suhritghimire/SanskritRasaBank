#!/usr/bin/env python3
"""
Restructure build_gradil_dataset.xlsx to include an 'extracted from' column
and reorder columns to: [verse_id, extracted from, shloka].
"""

import openpyxl
from openpyxl.styles import Font

XLSX_PATH = "/Users/suhritghimire/Downloads/NavaRasaBank-main/build_gradil_dataset.xlsx"

# Mapping of ID prefix to Work Name
MAPPING = {
    "KSak": "Abhijñānaśākuntalam",
    "HPri": "Priyadarśikā",
    "HDratn": "Ratnāvalī-nāṭikā",
    "KMgD": "Meghadūta",
    "Ragh": "Raghuvaṃśa",
    "BhKir": "Kirātārjunīya",
    "Bhall": "Bhallaṭaśataka",
    "Bhartri": "Śatakatraya",
    "MSubh": "Mahāsubhāṣitasaṅgraha",
    "SRK": "Subhāṣitaratnakoṣa",
    "KSS_Vet": "Kathāsaritsāgara (Vetala)",
    "Tantra1": "Tantrākhyāyika 1"
}

def get_work_name(verse_id):
    if not verse_id:
        return "Unknown"
    for prefix, name in MAPPING.items():
        if verse_id.startswith(prefix):
            return name
    return "Unknown"

print("Loading Excel file...")
wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.active

# Load all data
# Current columns: [verse_id, shloka]
data = list(ws.iter_rows(values_only=True))

if not data:
    print("No data found.")
    exit()

header = data[0]
rows = data[1:]

print("Restructuring data...")
new_header = ["verse_id", "extracted from", "shloka"]
new_rows = [new_header]

for r in rows:
    v_id = r[0]
    shloka = r[1]
    work_name = get_work_name(v_id)
    new_rows.append([v_id, work_name, shloka])

# Clear and rewrite
ws.delete_rows(1, ws.max_row)
for r in new_rows:
    ws.append(r)

# Format header
for cell in ws[1]:
    cell.font = Font(bold=True)

wb.save(XLSX_PATH)
print(f"Successfully restructured {XLSX_PATH}.")
print(f"Total rows: {len(new_rows)}")
