#!/usr/bin/env python3
"""
Refined Repair for Ratnāvalī (HDratn) verses.
- Backward-search from markers to get complete stanzas.
- Strict Sanskrit filtering (excludes verses with Prakrit telltales).
- Removal of title metadata and prose fragments.
- Split combined 'prose+verse' rows.
- Restore the 3-column format: [verse_id, extracted from, shloka].
"""

import re
import openpyxl
from indic_transliteration import sanscript
from indic_transliteration.sanscript import transliterate

XLSX_PATH = "/Users/suhritghimire/Downloads/NavaRasaBank-main/build_gradil_dataset.xlsx"
HTML_PATH = "/tmp/hdratn_u.htm"

PRAKRIT_TELLTALES = ['esā', 'khu', 'vaassa', 'pekkh', ' pautta', ' hiam', 'hadḍhī', ' jaṇa', ' edāe', ' kīsa', ' imiṇā', ' aaṃ', ' jjevva']

def is_sanskrit(text):
    text_lower = text.lower()
    # Check for Prakrit telltales
    for telltale in PRAKRIT_TELLTALES:
        if telltale in text_lower:
            return False
    # Reliable markers of Sanskrit in IAST
    if any(c in text for c in ['ś', 'ṣ', 'ṛ', 'ḥ', 'au', 'ai']):
        return True
    return False

def clean_and_translit(text):
    # Split after "tathā hi ।" if present
    if "tathā hi" in text.lower():
        parts = re.split(r'tathā hi\s*\|+', text, flags=re.IGNORECASE)
        if len(parts) > 1:
            text = parts[-1].strip()
            
    text = text.replace('||', '॥').replace('|', '।')
    dev = transliterate(text, sanscript.IAST, sanscript.DEVANAGARI)
    dev = dev.replace('।।', '॥')
    dev = re.sub(r'\s+', ' ', dev).strip()
    return dev

def extract_correct_ratnavali():
    with open(HTML_PATH, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    raw_lines = re.split(r'<BR>|\n', content)
    verses = []
    marker_pattern = re.compile(r'\|\|\s*(\d+)\s*\|\|\s*(\([\d.]+\))?')
    
    for i, line in enumerate(raw_lines):
        line = re.sub(r'<[^>]+>', '', line).strip()
        if not line: continue
            
        match = marker_pattern.search(line)
        if match:
            v_id_match = re.search(r'\(([\d.]+)\)', line)
            v_id_val = f"HDratn_{v_id_match.group(1)}" if v_id_match else f"HDratn_seq_{match.group(1)}"
            
            verse_lines = []
            last_line_text = marker_pattern.sub('', line).strip()
            if last_line_text:
                verse_lines.append(last_line_text)
            
            j = i - 1
            while j >= 0:
                prev_line = re.sub(r'<[^>]+>', '', raw_lines[j]).strip()
                if not prev_line:
                    j -= 1
                    continue
                
                if marker_pattern.search(prev_line): break
                if "Act" in prev_line or "Ratnāvalī" in prev_line: break
                
                speaker_match = re.match(r'^([A-Z][a-zāēīōūṣśṛḍṇ\s]+)\|+(.*)', prev_line)
                if speaker_match:
                    verse_part = speaker_match.group(2).strip()
                    if verse_part:
                        verse_lines.insert(0, verse_part)
                    break
                
                if prev_line.lower().startswith("api ca"):
                    verse_part = re.sub(r'^api ca\s*\|+', '', prev_line, flags=re.IGNORECASE).strip()
                    if verse_part:
                        verse_lines.insert(0, verse_part)
                    break
                
                if len(prev_line) > 250: break
                verse_lines.insert(0, prev_line)
                
                if len(verse_lines) >= 6: break
                j -= 1
            
            full_verse_text = " ".join(verse_lines)
            
            # Stricter checks: Metrical-looking (contains | or ||) and Sanskrit
            if " |" in full_verse_text or "||" in full_verse_text:
                if is_sanskrit(full_verse_text) and len(full_verse_text) > 40:
                    verses.append({
                        'id': v_id_val,
                        'text': clean_and_translit(full_verse_text)
                    })
    
    return verses

MAPPING = {
    "KSak": "Abhijñānaśākuntalam",
    "HPri": "Priyadarśikā",
    "HDratn": "Ratnāvalī-nāṭikā",
    "KMgD": "Meghadūta",
    "Ragh": "Raghuvaṃśa",
    "BhKir": "Kirātārjunīya",
    "Bhall": "Bhallaṭaśataka",
    "Bhartri": "Śatakatraya",
    "MSubh": "Mahāsubhāṣitasaṅgraha",
    "SRK": "Subhāṣitaratnakoṣa",
    "KSS_Vet": "Kathāsaritsāgara (Vetala)",
    "Tantra1": "Tantrākhyāyika 1"
}

def get_work_name(verse_id):
    if not verse_id: return "Unknown"
    for prefix, name in MAPPING.items():
        if verse_id.startswith(prefix):
            return name
    return "Unknown"

def repair_excel(new_verses):
    print(f"Repairing Excel with {len(new_verses)} verses...")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    
    old_data = list(ws.iter_rows(values_only=True))
    header = ["verse_id", "extracted from", "shloka"]
    
    cleaned_rows = []
    for row in old_data[1:]:
        if not row: continue
        v_id = row[0]
        # Detect shloka text - it might be in col 2 or 3 depending on previous state
        shloka = row[2] if len(row) >= 3 else row[1]
            
        if v_id and (str(v_id).startswith("HDratn_") or str(v_id).startswith("HDratn_seq_")):
            continue
        
        cleaned_rows.append([v_id, get_work_name(v_id), shloka])
    
    final_rows = [header]
    hpri_found = False
    hdratn_inserted = False
    
    for row in cleaned_rows:
        v_id = row[0]
        if v_id and str(v_id).startswith("HPri_"):
            hpri_found = True
        
        if hpri_found and not str(v_id).startswith("HPri_") and not hdratn_inserted:
            for nv in new_verses:
                final_rows.append([nv['id'], "Ratnāvalī-nāṭikā", nv['text']])
            hdratn_inserted = True
            
        final_rows.append(row)
        
    if not hdratn_inserted:
        for nv in new_verses:
            final_rows.append([nv['id'], "Ratnāvalī-nāṭikā", nv['text']])

    ws.delete_rows(1, ws.max_row)
    for r in final_rows:
        ws.append(r)
        
    wb.save(XLSX_PATH)
    print("Excel repaired with refined HDratn and 3-column format restored.")

if __name__ == "__main__":
    new_v = extract_correct_ratnavali()
    print(f"Extracted {len(new_v)} Sanskrit verses.")
    repair_excel(new_v)
