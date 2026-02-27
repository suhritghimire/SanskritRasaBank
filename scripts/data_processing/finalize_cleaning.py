#!/usr/bin/env python3
"""
Final detailed cleaning for BhKir and SRK/VidSrk.
- BhKir: replace '+' with ' '.
- SRK: remove numbering, citations, and metadata using regex.
- Ensures absolute 3-column format for all rows.
"""

import re
import openpyxl

XLSX_PATH = "/Users/suhritghimire/Downloads/NavaRasaBank-main/build_gradil_dataset.xlsx"

MAPPING = {
    "KSak": "Abhijñānaśākuntalam",
    "HPri": "Priyadarśikā",
    "HDratn": "Ratnāvalī-nāṭikā",
    "KMgD": "Meghadūta",
    "Ragh": "Raghuvaṃśa",
    "BhKir": "Kirātārjunīya",
    "Bhall": "Bhallaṭaśataka",
    "Bhartri": "Śatakatraya",
    "MSubh": "Mahāsubhāṣitasaṅgraha",
    "SRK": "Subhāṣitaratnakoṣa",
    "KSS_Vet": "Kathāsaritsāgara (Vetala)",
    "Tantra1": "Tantrākhyāyika 1"
}

def get_work_name(verse_id):
    if not verse_id: return "Unknown"
    for prefix, name in MAPPING.items():
        if verse_id.startswith(prefix):
            return name
    return "Unknown"

def clean_srk(text):
    # 1. Remove pure-number parentheses (Devanagari and Arabic)
    text = re.sub(r'\(\s*[०-९0-9]+\s*\)', '', text)
    # 1b. Remove pure-number with asterisks e.g. *(200)*
    text = re.sub(r'\*\(\s*[०-९0-9]+\s*\)\*', '', text)
    text = re.sub(r'\*\(\s*[०-९0-9]+\s*\)', '', text)
    
    # 2. Remove citation parentheses
    text = re.sub(r'\(\s*(Skmsa|Skms|स्क्म्स|स्प्द्शा|व्वाम्|सुभाष्|स्भ्सु)\.[^)]+\)', '', text)
    
    # 3. Remove trailing identifier blocks // ... //
    text = re.sub(r'//\s*[A-Za-z0-9_.]+\s*(?:\*\(\d+\))?\s*//\s*$', '', text)
    # Also catch simple // p.2 markers
    text = re.sub(r'//\s*p\.\s*\d+\s*,?', '', text)
    
    # 4. Remove author attribution marks like -- or -- some_name
    text = re.sub(r'--\s*[A-Za-zāēīōūṣśṛḍṇ\s]+(?:$|\n)', '', text)
    
    # 5. Cleanup artifacts
    text = re.sub(r'[\$\&\%]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    
    # Remove metadata lines that survived (e.g. "bhava-bhūteḥ (Skmsa.u.ka. 1256)")
    # If the text is just a name and no dandas, it's likely leftovers
    if '।' not in text and '॥' not in text and len(text) < 50:
        return ""
        
    return text

def clean_entire_dataset():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    
    # New rows buffer
    new_rows = [["verse_id", "extracted from", "shloka"]]
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]: continue
        
        v_id = str(row[0])
        work_name = get_work_name(v_id)
        # Use the right column for shloka
        shloka = row[2] if len(row) >= 3 else row[1]
        
        if not shloka: continue
        
        # Specific Work Cleaning
        if v_id.startswith("BhKir"):
            shloka = shloka.replace("+", " ")
            
        elif v_id.startswith("SRK"):
            shloka = clean_srk(shloka)
            if not shloka: continue # Skip if author name line was purged
            
        # Generic space cleanup
        shloka = re.sub(r'\s+', ' ', shloka).strip()
        
        new_rows.append([v_id, work_name, shloka])
        
    ws.delete_rows(1, ws.max_row)
    for r in new_rows:
        ws.append(r)
        
    # Bold header
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        
    wb.save(XLSX_PATH)
    print("Dataset cleaned and updated to 3-column format.")

if __name__ == "__main__":
    clean_entire_dataset()
