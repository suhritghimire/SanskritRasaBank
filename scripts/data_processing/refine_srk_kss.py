#!/usr/bin/env python3
"""
Refine SRK cleaning and regenerate KSS entries (Final Corrected Version).
- SRK: Remove only single-word leading attributions ending in ।.
- KSS: Merge consecutive half-line marked entries.
- Revert entire dataset to 2-column format.
"""

import re
import openpyxl
from indic_transliteration import sanscript
from indic_transliteration.sanscript import transliterate

XLSX_PATH = "/Users/suhritghimire/Downloads/NavaRasaBank-main/build_gradil_dataset.xlsx"
KSS_HTML_PATH = "/tmp/soksvppu.htm"

def clean_srk_text(text):
    # Remove leading attributions: e.g. "कस्यचित् ।", "अभिनन्दस्य ।"
    # Strict check: must be a single word (no internal spaces) before the first ।
    match = re.match(r'^([^।\s]+।)\s*', text)
    if match:
        attr = match.group(1)
        # Avoid stripping if it looks like a very short verse or anything else
        # Just to be safe, only strip if it's typical attribution length
        if len(attr) < 30:
            text = text[len(match.group(0)):].strip()
    
    # Remove citation parentheses with digits or keywords
    cit_keywords = ['वैस्', 'स्स्', 'स्क्म्स', 'स्प्द्शा', 'स्म्व्सू', 'स्व्सु', 'स्द्सा', 'Skmsa', 'Skms']
    cit_pattern = r'\(\s*(?:' + '|'.join(cit_keywords) + r'|[०-९0-9]|\.[^)]+)+\s*\)'
    text = re.sub(cit_pattern, '', text)
    
    # Remove remaining bracketed numeric markers like (१८८), (191), *(200)*
    text = re.sub(r'\*?\(\s*[०-९0-9]+\s*\)\*?', '', text)
    
    # Normalize spaces and remove leading ।
    text = re.sub(r'\s+', ' ', text).strip()
    if text.startswith('।'):
        text = text[1:].strip()
    
    return text

def strip_tags(s):
    s = re.sub(r"<[^>]+>", "", s)
    s = s.replace("&lt;", "<").replace("&gt;", ">").replace("&amp;", "&")
    return s.strip()

def extract_correct_kss():
    with open(KSS_HTML_PATH, "r", encoding="utf-8", errors="replace") as f:
        content = f.read()
    
    content = re.sub(r"<[Bb][Rr]\s*/?>", " <BR> ", content)
    raw_lines = re.split(r'\n', content)
    
    extracted_segments = []
    marker_pattern = re.compile(r"//\s*SoKss_[\d.,]+\s*\((Vet_([\d.]+))\)\s*//")
    
    buffer = []
    for line in raw_lines:
        line_clean = strip_tags(line).strip()
        if not line_clean: continue
            
        match = marker_pattern.search(line_clean)
        if match:
            v_id = match.group(1) # e.g. Vet_24.9
            verse_end_part = line_clean[:match.start()].strip()
            if verse_end_part:
                buffer.append(verse_end_part)
            
            full_text = " ".join(buffer)
            if "GRETIL" in full_text or "Somadeva" in full_text or "Vetala" in full_text:
                buffer = []
                continue
                
            extracted_segments.append({'id': v_id, 'text': full_text})
            buffer = []
        else:
            buffer.append(line_clean)
            
    merged_verses = []
    skip_next = False
    for i in range(len(extracted_segments)):
        if skip_next:
            skip_next = False
            continue
            
        curr = extracted_segments[i]
        
        if i + 1 < len(extracted_segments):
            nxt = extracted_segments[i+1]
            if '/' not in curr['text'] and '||' not in curr['text'] and '|' not in curr['text']:
                start_part = curr['id'].split('.')[-1]
                end_part = nxt['id'].split('.')[-1]
                prefix = ".".join(curr['id'].split('.')[:-1])
                merged_id = f"{prefix}.{start_part}-{end_part}"
                merged_text = curr['text'] + " " + nxt['text']
                merged_verses.append((merged_id, merged_text))
                skip_next = True
                continue
        
        merged_verses.append((curr['id'], curr['text']))

    final_verses = []
    for vid, txt in merged_verses:
        dev = transliterate(txt, sanscript.IAST, sanscript.DEVANAGARI)
        dev = dev.replace('||', '॥').replace('|', '।').replace('।।', '॥')
        dev = re.sub(r'\s+', ' ', dev).strip()
        final_verses.append((vid, dev))
        
    return final_verses

def run_refinement():
    print("Loading Excel...")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    
    old_data = list(ws.iter_rows(values_only=True))
    header = ["verse_id", "shloka"]
    new_rows = []
    
    print("Processing existing rows...")
    for row in old_data[1:]:
        if not row: continue
        v_id = str(row[0]) if row[0] else ""
        if len(row) >= 3:
            shloka = str(row[2]) if row[2] else ""
        else:
            shloka = str(row[1]) if row[1] else ""
            
        if v_id.startswith("KSS_Vet_") or v_id.startswith("SoKss_") or "(Vet_" in v_id or v_id.startswith("Vet_"):
            continue
            
        if v_id.startswith("SRK_") or v_id.startswith("VidSrk_"):
            shloka = clean_srk_text(shloka)
            if not shloka: continue

        new_rows.append([v_id, shloka])

    print("Regenerating KSS...")
    kss_verses = extract_correct_kss()
    for vid, text in kss_verses:
        new_rows.append([vid, text])

    print("Writing back to Excel...")
    ws.delete_rows(1, ws.max_row)
    ws.append(header)
    for r in new_rows:
        ws.append(r)
        
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        
    wb.save(XLSX_PATH)
    print("Repair complete.")

if __name__ == "__main__":
    run_refinement()
