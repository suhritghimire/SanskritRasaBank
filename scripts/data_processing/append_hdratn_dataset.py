#!/usr/bin/env python3
"""
Append Sanskrit verses from hdratn_u.htm (Harsha's Ratnāvalī-nāṭikā)
to the existing new_dataset_gretil.xlsx.

Format: verses span multiple plain HTML lines.
Last line of each verse ends with:  || N || (act.verse)
e.g. saindūrīkriyate ...kuṭṭimam || 11 || (1.11)<BR>

Verse ID used = act.verse form: HDratn_1.11
Prose lines are identified by patterns like:
  (N.N) SpeakerName |, iti prastāvanā, tataḥ praviśati, etc.
Prakrit dialogue lines (Naṭī, Vidūṣakaḥ) also excluded.
"""

import re
import openpyxl
from openpyxl.styles import Font

# ── 1. Read raw HTML ─────────────────────────────────────
with open("/tmp/hdratn_u.htm", encoding="utf-8", errors="replace") as f:
    raw_lines = f.read().splitlines()

print(f"Total raw lines: {len(raw_lines)}")

# ── Helper: strip HTML tags ───────────────────────────────
def strip_tags(s):
    s = re.sub(r"<[^>]+>", "", s)
    s = s.replace("&lt;","<").replace("&gt;",">").replace("&amp;","&")
    return s.strip()

# ── 2. Patterns ───────────────────────────────────────────
# Verse-closing line: || N || (act.verse) OR just || N ||
# We capture the act.verse tuple like (1.11) as the ID
ID_PAT = re.compile(
    r"\|\|\s*(\d+)\s*\|\|(?:\s*\((\d+)\.(\d+)\))?",
)

# Prose line indicators — lines we SKIP as non-verse
PROSE_PAT = re.compile(
    r"^\s*("
    r"\(\d+\.\d+\)"          # e.g. (1.5) Sūtradhāraḥ |
    r"|tataḥ praviśat"
    r"|iti prastāvanā"
    r"|iti viṣkambhakaḥ"
    r"|iti niṣkrāntaḥ"
    r"|iti niṣkrāntau"
    r"|nāndyante"
    r"|nepathye"
    r"|api ca"               # standalone "api ca |" prose connector
    r"|nāndī"
    r"|Harsadeva"
    r"|Based on"
    r"|THIS GRETIL"
    r"|Copyright"
    r"|Text converted"
    r"|Unless"
    r"|description:"
    r"|long [aiuAIU]"
    r"|vocalic|retroflex|palatal|velar|anusvara|visarga"
    r"|For a comp|For further"
    r")",
    re.IGNORECASE,
)

# Lines that start with a speaker tag like  "Rājā |"  or  "(1.5) Rājā |"
SPEAKER_PAT = re.compile(
    r"^\s*(\(\d+\.\d+\)\s*)?[A-ZĀĪŪ][a-zāīū]+[āīūḥṃ]?\s*\|",
)

SKIP_METADATA = re.compile(
    r"^\s*(\[p\.\s*\d+\]|http|GRETIL|Input|Harsadeva|Based|3\.|verbesserte"
    r"|herausgegeben|Copyright|THIS|Text|Unless|description:|long |vocalic"
    r"|retroflex|palatal|velar|anusvara|visarga|For a|___)",
    re.IGNORECASE,
)

records  = []
buf      = []
cur_id   = ""
act_num  = 1   # track current act for IDs without explicit act number

for raw_line in raw_lines:
    clean = strip_tags(raw_line).strip()

    # Skip blank lines — don't flush on them (verses span blanks in this file)
    if not clean:
        continue

    # Skip metadata / editorial headers
    if SKIP_METADATA.match(clean):
        continue

    # Skip prose / speaker lines
    if PROSE_PAT.match(clean) or SPEAKER_PAT.match(clean):
        if buf:
            buf = []   # abandon incomplete buffer — it mixed with prose
        continue

    # Also skip Prakrit lines that look like Middle Indian
    # (rough heuristic: contain ṇ/ṃ heavy patterns without Sanskrit sandhi,
    #  OR start with lowercase characters followed by Prakrit suffixes -ao, -aṃ)
    # We'll just check for the common Prakrit tag patterns
    if re.match(r"^\(1\.\d+\)", clean):
        if buf:
            buf = []
        continue

    # Does this line carry the closing || N || marker?
    id_m = ID_PAT.search(clean)
    if id_m:
        verse_num = id_m.group(1)
        if id_m.group(2) and id_m.group(3):
            act_num = id_m.group(2)
            within  = id_m.group(3)
            cur_id  = f"HDratn_{act_num}.{within}"
        else:
            cur_id  = f"HDratn_{act_num}.{verse_num}"

        # Strip the || N || (act.verse) tag from the content
        content = ID_PAT.sub("", clean).strip().strip(" /|.")
        if content:
            buf.append(content)

        # Flush verse
        verse = " ".join(buf)
        verse = re.sub(r"\s+", " ", verse).strip().strip(" /|.")
        verse = re.sub(r"\[p\.\s*\d+\]", "", verse).strip()
        if len(verse) > 8:
            records.append((cur_id, verse))
        buf    = []
        cur_id = ""
        continue

    # Plain verse content — accumulate
    clean = re.sub(r"\[p\.\s*\d+\]", "", clean).strip()
    # Skip lines that are clearly Prakrit (simple heuristic: end with -ao, -aṃ, -aī with no ||)
    if re.search(r"(ao|aṃ|ḍhi|ṇai|ṇo)\s*$", clean) and not id_m:
        # Likely Prakrit — skip
        if buf:
            buf = []
        continue

    if clean:
        buf.append(clean)

print(f"Extracted {len(records)} HDratn verses")

# ── 3. Transliterate ──────────────────────────────────────
def transliterate(text):
    try:
        from indic_transliteration import sanscript
        from indic_transliteration.sanscript import transliterate as tr
        return tr(text, sanscript.IAST, sanscript.DEVANAGARI)
    except Exception:
        pass
    TABLE = [
        ("ā","आ"),("Ā","आ"),("ī","ई"),("Ī","ई"),("ū","ऊ"),("Ū","ऊ"),
        ("ṝ","ॠ"),("ṛ","ऋ"),("Ṛ","ऋ"),("ḷ","ऌ"),("ḹ","ॡ"),
        ("ai","ऐ"),("au","औ"),("e","ए"),("o","ओ"),
        ("a","अ"),("i","इ"),("u","उ"),
        ("ṃ","ं"),("ḥ","ः"),
        ("kh","ख"),("gh","घ"),("ch","छ"),("jh","झ"),
        ("ṭh","ठ"),("ḍh","ढ"),("th","थ"),("dh","ध"),
        ("ph","फ"),("bh","भ"),
        ("k","क"),("g","ग"),("ṅ","ङ"),
        ("c","च"),("j","ज"),("ñ","ञ"),
        ("ṭ","ट"),("ḍ","ड"),("ṇ","ण"),
        ("t","त"),("d","द"),("n","न"),
        ("p","प"),("b","ब"),("m","म"),
        ("y","य"),("r","र"),("l","ल"),("v","व"),
        ("ś","श"),("Ś","श"),("ṣ","ष"),("Ṣ","ष"),
        ("s","स"),("h","ह"),("ḻ","ळ"),
        ("||","॥"),("|","।"),
    ]
    out, i = [], 0
    while i < len(text):
        for src, dst in TABLE:
            if text[i:i+len(src)] == src:
                out.append(dst); i += len(src); break
        else:
            out.append(text[i]); i += 1
    return "".join(out)

print("Transliterating …")
final = [(vid, transliterate(v)) for vid, v in records]
print(f"Done — {len(final)} rows to append")

# ── 4. Append to existing Excel ───────────────────────────
XLSX = "/Users/suhritghimire/Downloads/NavaRasaBank-main/new_dataset_gretil.xlsx"
wb   = openpyxl.load_workbook(XLSX)
ws   = wb.active

before = ws.max_row
for vid, dev in final:
    ws.append([dev, vid])
wb.save(XLSX)

print(f"\n✅  Saved → {XLSX}")
print(f"   Previous data rows : {before - 1}")
print(f"   Appended HDratn rows: {len(final)}")
print(f"   Total data rows now : {ws.max_row - 1}")
print("\nSample HDratn rows:")
for vid, dev in final[:5]:
    print(f"  [{vid}]  {dev[:90]}")
